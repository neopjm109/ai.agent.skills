#!/usr/bin/env python3
"""Dump every sheet of an .xlsx workbook as tab-separated rows.

Prints each sheet as a `=== SHEET: name ===` block of tab-joined rows. Prefers
openpyxl; falls back to parsing the workbook XML (an .xlsx is a ZIP) with the
standard library only, so the sole requirement is python3.

Usage: python3 scripts/extract.py <file.xlsx>
"""
import sys


def with_openpyxl(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        print(f"=== SHEET: {ws.title} ===")
        for row in ws.iter_rows(values_only=True):
            print("\t".join("" if c is None else str(c) for c in row))


def with_stdlib(path):
    import re
    import zipfile
    from xml.etree import ElementTree as ET

    NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    PKG_REL = "{http://schemas.openxmlformats.org/package/2006/relationships}"

    def col_index(ref):
        letters = re.match(r"[A-Z]+", ref).group(0)
        n = 0
        for ch in letters:
            n = n * 26 + (ord(ch) - ord("A") + 1)
        return n - 1

    with zipfile.ZipFile(path) as z:
        names = set(z.namelist())

        shared = []
        if "xl/sharedStrings.xml" in names:
            sroot = ET.fromstring(z.read("xl/sharedStrings.xml"))
            for si in sroot.findall(f"{NS}si"):
                shared.append("".join(t.text or "" for t in si.iter(f"{NS}t")))

        # rId -> worksheet target
        targets = {}
        if "xl/_rels/workbook.xml.rels" in names:
            rroot = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
            for rel in rroot.findall(f"{PKG_REL}Relationship"):
                tgt = rel.get("Target")
                targets[rel.get("Id")] = tgt if tgt.startswith("xl/") else f"xl/{tgt}"

        wbroot = ET.fromstring(z.read("xl/workbook.xml"))
        for sheet in wbroot.find(f"{NS}sheets"):
            title = sheet.get("name")
            path_in_zip = targets.get(sheet.get(f"{REL}id"))
            print(f"=== SHEET: {title} ===")
            if not path_in_zip or path_in_zip not in names:
                continue
            wsroot = ET.fromstring(z.read(path_in_zip))
            for row in wsroot.iter(f"{NS}row"):
                cells, maxc = {}, -1
                for c in row.findall(f"{NS}c"):
                    ref = c.get("r") or "A"
                    idx = col_index(ref)
                    maxc = max(maxc, idx)
                    v = c.find(f"{NS}v")
                    if c.get("t") == "s" and v is not None:
                        cells[idx] = shared[int(v.text)]
                    elif c.get("t") == "inlineStr":
                        cells[idx] = "".join(t.text or "" for t in c.iter(f"{NS}t"))
                    elif v is not None:
                        cells[idx] = v.text or ""
                print("\t".join(cells.get(i, "") for i in range(maxc + 1)))


def main():
    if len(sys.argv) != 2:
        sys.exit("usage: extract.py <file.xlsx>")
    path = sys.argv[1]
    try:
        with_openpyxl(path)
    except ImportError:
        with_stdlib(path)


if __name__ == "__main__":
    main()
